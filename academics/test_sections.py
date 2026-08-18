"""
Sections: a subdivision of one batch, and the rules that keep a student in one
of their own.

The test worth reading is `test_section_a_of_two_batches_are_two_sections`.
Everything else follows from it — a section is scoped to a cohort, so "A" is
not one thing across the college, and every path that writes one has to check
it belongs to the batch being written.
"""
import io
import json

from django.test import RequestFactory, TestCase
from django.urls import reverse

from academics import sections as svc
from academics import views
from academics.importer import ROSTER_COLUMNS, import_students, read_rows
from academics.models import (
    Batch,
    Department,
    Enrollment,
    Section,
    StudentProfile,
    Subject,
)
from accounts.models import Institute, User


def sheet(rows, headers=None):
    """An in-memory .xlsx the importer can read."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(headers or ROSTER_COLUMNS)
    for row in rows:
        ws.append(row)
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


class SectionFixture(TestCase):
    def setUp(self):
        self.institute = Institute.objects.create(name="Acme", code="ACME",
                                                  email="a@a.edu")
        self.dept = Department.objects.create(institute=self.institute,
                                              name="Computer Science",
                                              code="CSE")
        self.batch = Batch.objects.create(department=self.dept,
                                          label="2022-26", start_year=2022,
                                          end_year=2026)
        self.other_batch = Batch.objects.create(department=self.dept,
                                                label="2023-27",
                                                start_year=2023, end_year=2027)
        self.subject = Subject.objects.create(department=self.dept, code="DSA",
                                              name="Data Structures")
        self.head = User.objects.create_user(
            email="head@a.edu", password="Str0ngPass!23", role=User.Role.HEAD,
            institute=self.institute, registration_completed=True)

    def student(self, email="s@a.edu", batch=None, section=None):
        user = User.objects.create_user(
            email=email, password="Str0ngPass!23", role=User.Role.STUDENT,
            institute=self.institute, department=self.dept,
            full_name="Asha Rao", registration_completed=True)
        profile = StudentProfile.objects.create(
            user=user, department=self.dept, batch=batch or self.batch,
            section=section, class_roll="01",
            guardian_mobile="+919812345670")
        Enrollment.objects.create(student=profile, subject=self.subject,
                                  is_active=True)
        return profile

    def import_sheet(self, rows, **kwargs):
        parsed, error = read_rows(sheet(rows))
        self.assertIsNone(error, error)
        # Institute-scoped: the department is a column in the sheet now.
        return import_students(parsed, self.institute, self.head, **kwargs)


class ScopeTests(SectionFixture):
    def test_section_a_of_two_batches_are_two_sections(self):
        """
        The premise of the whole design. Two cohorts each have an "A", and they
        are different groups of people — one row meaning both would let the
        first cohort to graduate take the other's students with it.
        """
        a1, _ = svc.resolve(self.batch, "A", create=True)
        a2, _ = svc.resolve(self.other_batch, "A", create=True)
        self.assertNotEqual(a1.pk, a2.pk)
        self.assertEqual(Section.objects.filter(name="A").count(), 2)

    def test_one_batch_cannot_hold_the_same_section_twice(self):
        first, made_first = svc.resolve(self.batch, "A", create=True)
        again, made_again = svc.resolve(self.batch, "A", create=True)
        self.assertEqual(first.pk, again.pk)
        self.assertTrue(made_first)
        self.assertFalse(made_again)

    def test_names_are_normalised_so_one_section_has_one_spelling(self):
        first, _ = svc.resolve(self.batch, "A", create=True)
        for spelling in ("a", " A ", "a "):
            found, made = svc.resolve(self.batch, spelling, create=True)
            self.assertEqual(found.pk, first.pk, spelling)
            self.assertFalse(made, spelling)
        self.assertEqual(Section.objects.count(), 1)

    def test_a_blank_name_is_no_section_rather_than_an_error(self):
        self.assertEqual(svc.resolve(self.batch, "  "), (None, False))

    def test_resolving_without_create_refuses_an_unknown_section(self):
        with self.assertRaises(svc.SectionError):
            svc.resolve(self.batch, "Z")

    def test_a_section_of_another_batch_is_refused(self):
        foreign, _ = svc.resolve(self.other_batch, "A", create=True)
        with self.assertRaises(svc.SectionError):
            svc.assert_in_batch(foreign, self.batch)

    def test_deleting_a_section_does_not_delete_its_students(self):
        """`SET_NULL`. They become unsectioned, which every screen renders."""
        section, _ = svc.resolve(self.batch, "A", create=True)
        profile = self.student(section=section)
        section.delete()
        profile.refresh_from_db()
        self.assertIsNone(profile.section_id)
        self.assertTrue(StudentProfile.objects.filter(pk=profile.pk).exists())


class ImportTests(SectionFixture):
    def row(self, email="s@a.edu", section="A", batch="2022-26", dept="CSE"):
        """One row in `ROSTER_COLUMNS` order — Department is the third cell."""
        return [email, "Asha Rao", dept, "01", "CSE22001", batch, section,
                "DSA", "+919812345670", "Mr Rao", "", "9876543210", ""]

    def test_a_section_named_in_the_sheet_is_created_and_assigned(self):
        job = self.import_sheet([self.row()])
        self.assertEqual(job.error_count, 0, job.report)
        profile = StudentProfile.objects.get()
        self.assertEqual(profile.section.name, "A")
        self.assertEqual(profile.section.batch, self.batch)

    def test_two_rows_naming_one_new_section_create_it_once(self):
        """
        The second row must find the first rather than trip the uniqueness
        constraint and fail the whole file.
        """
        job = self.import_sheet([self.row("a@a.edu"), self.row("b@a.edu")])
        self.assertEqual(job.error_count, 0, job.report)
        self.assertEqual(Section.objects.filter(batch=self.batch).count(), 1)

    def test_the_sections_it_created_are_named_in_the_report(self):
        """
        Named, not counted — "2022-26 · AA" is the line that tells somebody
        they typed the section twice.
        """
        job = self.import_sheet([self.row("a@a.edu", section="A"),
                                 self.row("b@a.edu", section="B")])
        self.assertEqual(sorted(job.report["sections_created"]),
                         ["2022-26 · A", "2022-26 · B"])

    def test_an_existing_section_is_reused_not_reported_as_created(self):
        svc.resolve(self.batch, "A", create=True)
        job = self.import_sheet([self.row()])
        self.assertEqual(job.report["sections_created"], [])
        self.assertEqual(Section.objects.count(), 1)

    def test_a_blank_section_leaves_an_existing_one_alone(self):
        """Blank means "not changing this", as it does for every other column."""
        section, _ = svc.resolve(self.batch, "A", create=True)
        profile = self.student(section=section)
        self.import_sheet([self.row(email=profile.user.email, section="")])
        profile.refresh_from_db()
        self.assertEqual(profile.section_id, section.pk)

    def test_a_single_dash_clears_the_section(self):
        section, _ = svc.resolve(self.batch, "A", create=True)
        profile = self.student(section=section)
        self.import_sheet([self.row(email=profile.user.email, section="-")])
        profile.refresh_from_db()
        self.assertIsNone(profile.section_id)

    def test_the_column_is_optional(self):
        """A college that does not divide its cohorts omits it entirely."""
        headers = [h for h in ROSTER_COLUMNS if h != "Section"]
        row = [c for i, c in enumerate(self.row())
               if ROSTER_COLUMNS[i] != "Section"]
        parsed, error = read_rows(sheet([row], headers))
        self.assertIsNone(error, error)
        job = import_students(parsed, self.dept, self.head)
        self.assertEqual(job.error_count, 0, job.report)
        self.assertIsNone(StudentProfile.objects.get().section_id)

    def test_moving_a_student_to_another_batch_does_not_carry_the_section(self):
        """
        A student listed under another cohort's section is a quiet error in two
        rosters at once. Unsectioned is an ordinary state; this is not.
        """
        section, _ = svc.resolve(self.batch, "A", create=True)
        profile = self.student(section=section)
        self.import_sheet([self.row(email=profile.user.email, section="",
                                    batch="2023-27")])
        profile.refresh_from_db()
        self.assertEqual(profile.batch, self.other_batch)
        self.assertIsNone(profile.section_id)

    def test_a_row_moving_batch_and_naming_a_section_lands_in_the_new_one(self):
        profile = self.student()
        self.import_sheet([self.row(email=profile.user.email, section="A",
                                    batch="2023-27")])
        profile.refresh_from_db()
        self.assertEqual(profile.section.batch, self.other_batch)


class DepartmentColumnTests(SectionFixture):
    """
    The department comes from the row, not from a dropdown.

    One file can carry the whole college; the server still narrows what a row
    may name to the departments the uploader manages.
    """

    def setUp(self):
        super().setUp()
        self.other_dept = Department.objects.create(
            institute=self.institute, name="Electronics", code="ECE")
        Batch.objects.create(department=self.other_dept, label="2022-26",
                             start_year=2022, end_year=2026)
        Subject.objects.create(department=self.other_dept, code="SIG",
                               name="Signals")

    def row(self, email, dept, subjects="DSA"):
        return [email, "Asha Rao", dept, "01", "", "2022-26", "", subjects,
                "+919812345670", "Mr Rao", "", "9876543210", ""]

    def test_one_file_can_reach_several_departments(self):
        job = self.import_sheet([self.row("a@a.edu", "CSE"),
                                 self.row("b@a.edu", "ECE", subjects="SIG")])
        self.assertEqual(job.error_count, 0, job.report)
        self.assertEqual(
            StudentProfile.objects.get(user__email="a@a.edu").department,
            self.dept)
        self.assertEqual(
            StudentProfile.objects.get(user__email="b@a.edu").department,
            self.other_dept)

    def test_the_department_matches_on_name_as_well_as_code(self):
        job = self.import_sheet([self.row("a@a.edu", "Computer Science")])
        self.assertEqual(job.error_count, 0, job.report)
        self.assertEqual(
            StudentProfile.objects.get(user__email="a@a.edu").department,
            self.dept)

    def test_an_unknown_department_is_an_error_not_a_new_department(self):
        job = self.import_sheet([self.row("a@a.edu", "MECH")])
        self.assertEqual(job.error_count, 1)
        self.assertIn("MECH", job.report["rows"][0]["messages"][0])
        self.assertFalse(Department.objects.filter(code="MECH").exists())

    def test_a_new_student_without_a_department_is_refused(self):
        """There is no dropdown to fall back on any more."""
        job = self.import_sheet([self.row("a@a.edu", "")])
        self.assertEqual(job.error_count, 1)
        self.assertIn("department is blank",
                      " ".join(job.report["rows"][0]["messages"]))

    def test_an_existing_student_may_leave_it_blank(self):
        profile = self.student()
        job = self.import_sheet([self.row(profile.user.email, "")])
        self.assertEqual(job.error_count, 0, job.report)
        profile.refresh_from_db()
        self.assertEqual(profile.department, self.dept)

    def test_a_hod_cannot_file_students_into_another_department(self):
        """
        The guard the dropdown used to provide for free. Without it a Department
        column would let a HoD put students anywhere in the college.
        """
        parsed, _ = read_rows(sheet([self.row("a@a.edu", "ECE",
                                              subjects="SIG")]))
        job = import_students(parsed, self.institute, self.head,
                              allowed_departments=Department.objects.filter(
                                  pk=self.dept.pk))
        self.assertEqual(job.error_count, 1)
        self.assertIn("not one you manage",
                      job.report["rows"][0]["messages"][0])

    def test_subjects_are_resolved_inside_the_rows_own_department(self):
        """
        "SIG" exists in ECE and not in CSE. A row naming CSE must not find it.
        """
        job = self.import_sheet([self.row("a@a.edu", "CSE", subjects="SIG")])
        self.assertEqual(job.error_count, 1)
        self.assertIn("unknown subject(s) in CSE",
                      job.report["rows"][0]["messages"][0])

    def test_the_job_records_the_institute_and_names_the_departments(self):
        job = self.import_sheet([self.row("a@a.edu", "CSE"),
                                 self.row("b@a.edu", "ECE", subjects="SIG")])
        self.assertEqual(job.institute, self.institute)
        self.assertIsNone(job.department_id)
        self.assertEqual(job.report["departments"], ["CSE", "ECE"])

    def test_moving_a_student_between_departments_needs_a_batch_there(self):
        """
        Their old batch belongs to the department they are leaving. Refusing
        beats guessing — picking a cohort for somebody puts them in a year
        nobody chose.
        """
        profile = self.student()
        row = self.row(profile.user.email, "ECE", subjects="SIG")
        row[5] = ""                                    # no batch named
        job = self.import_sheet([row])
        self.assertEqual(job.error_count, 1)
        self.assertIn("has to name a batch in ECE",
                      job.report["rows"][0]["messages"][0])

    def test_discipline_is_no_longer_a_column(self):
        from academics.importer import ROSTER_COLUMNS

        self.assertNotIn("Discipline", ROSTER_COLUMNS)
        self.assertIn("Department", ROSTER_COLUMNS)

    def test_an_old_sheet_with_a_discipline_column_still_parses(self):
        """The alias is kept so files exported before this change still load."""
        headers = ROSTER_COLUMNS + ["Discipline"]
        parsed, error = read_rows(
            sheet([self.row("a@a.edu", "CSE") + ["Engineering"]], headers))
        self.assertIsNone(error, error)
        job = import_students(parsed, self.institute, self.head)
        self.assertEqual(job.error_count, 0, job.report)


class RoundTripTests(SectionFixture):
    def test_the_template_and_the_export_agree_on_the_column(self):
        """
        One list, `ROSTER_COLUMNS`, so a round trip lines up cell for cell.
        The two literal lists this replaced drifted the moment either gained a
        column.
        """
        from academics.importer import build_roster_workbook, build_template_workbook
        from openpyxl import load_workbook

        template = load_workbook(build_template_workbook(self.dept))["Students"]
        export = load_workbook(build_roster_workbook([]))["Students"]
        headers_t = [c.value for c in template[1]]
        headers_e = [c.value for c in export[1]]
        self.assertIn("Section", headers_t)
        self.assertEqual(headers_t.index("Section"),
                         headers_e.index("Section"))

    def test_an_exported_roster_carries_the_section(self):
        from academics.importer import build_roster_workbook
        from openpyxl import load_workbook

        section, _ = svc.resolve(self.batch, "A", create=True)
        profile = self.student(section=section)
        qs = StudentProfile.objects.select_related(
            "user", "batch", "section", "department", "department__institute")
        ws = load_workbook(build_roster_workbook(qs))["Students"]
        headers = [c.value for c in ws[1]]
        values = [c.value for c in ws[2]]
        self.assertEqual(values[headers.index("Section")], "A")
        self.assertEqual(values[headers.index("Email")], profile.user.email)

    def test_an_unsectioned_student_exports_a_blank_not_a_crash(self):
        from academics.importer import build_roster_workbook
        from openpyxl import load_workbook

        self.student()
        qs = StudentProfile.objects.select_related(
            "user", "batch", "section", "department", "department__institute")
        ws = load_workbook(build_roster_workbook(qs))["Students"]
        headers = [c.value for c in ws[1]]
        values = [c.value for c in ws[2]]
        self.assertIn(values[headers.index("Section")], ("", None))


class EditModalTests(SectionFixture):
    def call(self, view, user, pk=None, **data):
        request = RequestFactory().post("/", data)
        request.user = user
        request.session = self.client.session
        request.META["HTTP_X_REQUESTED_WITH"] = "XMLHttpRequest"
        return view(request, pk=pk) if pk is not None else view(request)

    def save(self, profile, **extra):
        data = {"full_name": "Asha Rao", "class_roll": "01",
                "batch_id": str(profile.batch_id),
                "guardian_mobile": "+919812345670"}
        data.update(extra)
        return self.call(views.api_student_save, self.head, pk=profile.pk,
                         **data)

    def test_a_student_can_be_reassigned_to_another_section(self):
        a, _ = svc.resolve(self.batch, "A", create=True)
        b, _ = svc.resolve(self.batch, "B", create=True)
        profile = self.student(section=a)
        response = self.save(profile, section_id=str(b.pk))
        self.assertTrue(json.loads(response.content)["success"],
                        response.content)
        profile.refresh_from_db()
        self.assertEqual(profile.section_id, b.pk)

    def test_a_student_can_be_taken_out_of_a_section(self):
        a, _ = svc.resolve(self.batch, "A", create=True)
        profile = self.student(section=a)
        self.save(profile, section_id="")
        profile.refresh_from_db()
        self.assertIsNone(profile.section_id)

    def test_another_batchs_section_is_refused(self):
        foreign, _ = svc.resolve(self.other_batch, "A", create=True)
        profile = self.student()
        response = self.save(profile, section_id=str(foreign.pk))
        body = json.loads(response.content)
        self.assertFalse(body["success"])
        self.assertIn("section_id", body["errors"])
        profile.refresh_from_db()
        self.assertIsNone(profile.section_id)

    def test_moving_batch_and_section_together_checks_the_new_batch(self):
        """
        Resolved against the batch being saved, not the one being left — the
        other way round lets a mismatch through in exactly this case.
        """
        target, _ = svc.resolve(self.other_batch, "A", create=True)
        profile = self.student()
        response = self.save(profile, batch_id=str(self.other_batch.pk),
                             section_id=str(target.pk))
        self.assertTrue(json.loads(response.content)["success"],
                        response.content)
        profile.refresh_from_db()
        self.assertEqual(profile.batch, self.other_batch)
        self.assertEqual(profile.section_id, target.pk)


class TableTests(SectionFixture):
    def rows(self, **params):
        self.client.force_login(self.head)
        return self.client.get(
            reverse("academics:api_students"), params,
            HTTP_X_REQUESTED_WITH="XMLHttpRequest").json()["data"]["rows"]

    def test_the_row_carries_the_section_name_and_id(self):
        section, _ = svc.resolve(self.batch, "A", create=True)
        self.student(section=section)
        row = self.rows()[0]
        self.assertEqual(row["section"], "A")
        self.assertEqual(str(row["section_id"]), str(section.pk))

    def test_an_unsectioned_student_reports_an_empty_section(self):
        self.student()
        row = self.rows()[0]
        self.assertEqual(row["section"], "")
        self.assertIsNone(row["section_id"])

    def test_filtering_by_section(self):
        a, _ = svc.resolve(self.batch, "A", create=True)
        self.student("a@a.edu", section=a)
        self.student("b@a.edu")
        rows = self.rows(section=str(a.pk))
        self.assertEqual([r["email"] for r in rows], ["a@a.edu"])

    def test_filtering_by_not_in_a_section(self):
        """The question a head asks straight after an import."""
        a, _ = svc.resolve(self.batch, "A", create=True)
        self.student("a@a.edu", section=a)
        self.student("b@a.edu")
        rows = self.rows(section="none")
        self.assertEqual([r["email"] for r in rows], ["b@a.edu"])

    def test_the_page_offers_live_sections_in_the_filter(self):
        live, _ = svc.resolve(self.batch, "A", create=True)
        retired, _ = svc.resolve(self.batch, "B", create=True)
        retired.is_active = False
        retired.save()
        self.client.force_login(self.head)
        offered = {s.pk for s in self.client.get(
            reverse("academics:students")).context["filter_sections"]}
        self.assertIn(live.pk, offered)
        self.assertNotIn(retired.pk, offered)
