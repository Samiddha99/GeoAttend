"""
Student roster import from .xlsx / .csv.

Columns (header row, case/spacing insensitive, extras ignored):
    email | name | class roll | batch | subjects | guardian mobile
    mobile | exam roll | guardian name | guardian email        (optional)

`subjects` accepts a comma / semicolon / pipe separated list of subject codes
or subject names, e.g.  "DSA, DBMS, AI".

The sheet is an **upsert keyed on email**. A row whose email already exists
updates that student instead of creating a second one, which is what makes bulk
edits possible: upload a two-column sheet of email + guardian mobile and only
guardian mobiles change.

That only works if a blank cell means "leave this alone" rather than "set this
to empty" — otherwise a partial sheet would silently wipe names, batches and
enrolments. So blanks are preserved for students who already exist, and the
required columns are only enforced for students being created. To actually
clear a field, put a single "-" in the cell.
"""
import csv
import io
import re

from django.db import transaction

from accounts.emails import send_invitation
from accounts.services import invite_user
from core.utils import normalise_email, parse_batch_label

from . import sections
from .models import Batch, Enrollment, ImportJob, StudentProfile, Subject

HEADER_ALIASES = {
    "name": {"name", "student name", "studentname", "full name", "fullname", "student"},
    "email": {"email", "email id", "emailid", "e-mail", "mail", "email address"},
    "mobile": {"mobile", "mobile number", "phone", "phone number", "contact", "contact number"},
    "batch": {"batch", "session", "year", "batch year", "admission batch"},
    "subjects": {"subjects", "subject", "subjects enrolled", "enrolled subjects",
                 "subject enrolled", "courses", "papers"},
    # Two rolls. Bare "roll"/"roll no" stays the class roll, since that is what
    # existing sheets used; registration/university numbers are exam rolls.
    "class_roll": {"class roll", "class roll no", "class roll no.", "class roll number",
                   "roll", "roll no", "roll no.", "roll number", "rollno", "class no",
                   "class number"},
    # Which section of the batch. Optional everywhere: a college that does not
    # divide its cohorts leaves the column out, and one that does can fill it
    # in later without re-uploading anything else.
    "section": {"section", "sec", "section name", "class section", "division",
                "div", "group"},
    "exam_roll": {"exam roll", "exam roll no", "exam roll no.", "exam roll number",
                  "examroll", "registration no", "registration number", "reg no",
                  "university roll", "university roll no", "admit card no", "exam no"},
    "guardian_email": {"guardian email", "parent email", "parent mail",
                       "guardian mail", "guardian email id"},
    "guardian_mobile": {"guardian mobile", "guardian mobile number", "guardian number",
                        "guardian phone", "guardian contact", "parent mobile",
                        "parent mobile number", "parent number", "parent phone",
                        "parent contact", "guardian whatsapp", "whatsapp number",
                        "guardian", "parent"},
    "guardian_name": {"guardian name", "parent name", "father name", "mother name",
                      "guardian/parent name"},
    # **An instruction, not a check.** The department used to be chosen once
    # for the whole file, from a dropdown, which meant a college with six
    # departments uploaded six spreadsheets. It is a column now: one file can
    # carry the whole institute, and each row says where its student belongs.
    #
    # Matched on code or name, the same way subjects are — "CSE" and "Computer
    # Science" both find the same department, because a registrar's export
    # spells it whichever way their other system does.
    "department": {"department", "department name", "department code", "dept",
                   "dept name", "dept code", "branch"},
    # A *check*, not an instruction. A student's institute follows from the
    # department their row names, so a cell cannot move them — but a sheet
    # naming a different institute is almost always a sheet uploaded to the
    # wrong account, which is worth catching before it lands. Optional: omit it
    # and nothing is checked.
    "institute": {"institute", "institute name", "college", "college name",
                  "institution"},
    # Kept, deliberately unused. Discipline was a checked column and is not one
    # any more: it follows from the department, which the row now names, so
    # there is nothing left for it to catch that the department does not. The
    # alias stays so that a sheet exported before this change still parses —
    # the column is simply ignored rather than reported as unknown.
    "discipline": {"discipline", "stream", "faculty", "branch of study"},
}

# Email is the key the upsert matches on, so it is the only header the file
# genuinely cannot do without. Everything else is validated per row, and only
# for students that do not exist yet — see NEW_STUDENT_REQUIRED.
REQUIRED_COLUMNS = ("email",)

# Enforced only when creating. An existing student may omit any of these.
#
# `department` joined this list when it stopped being a dropdown: a new student
# has to belong somewhere, and there is no longer a single answer to fall back
# on. An *existing* student may leave it blank, which keeps theirs — the same
# rule every other column follows.
NEW_STUDENT_REQUIRED = ("name", "department", "class_roll", "batch",
                        "subjects", "guardian_mobile")

# The one column order, used by the template the head downloads *and* by the
# roster export they upload back. Two literal lists drifted the moment either
# gained a column — the export still ended at "Mobile Number" while the
# template had grown Institute and Discipline — so a round trip put values
# under the wrong headings. One list cannot drift from itself.
#
# `Status` is deliberately outside it: it is an export-only note, and it sits
# after the shared columns so the two files line up cell for cell until it.
ROSTER_COLUMNS = [
    "Email", "Name", "Department", "Class Roll", "Exam Roll", "Batch",
    "Section", "Subjects Enrolled", "Guardian Mobile", "Guardian Name",
    "Guardian Email", "Mobile Number", "Institute",
]
EXPORT_ONLY_COLUMNS = ["Status"]


def _same(a, b):
    """Loose string equality — case and surrounding space do not matter."""
    return (a or "").strip().casefold() == (b or "").strip().casefold()


# Put this in a cell to blank a value that is currently set. Without it there
# would be no way to erase a guardian email, because blank means "unchanged".
CLEAR_TOKEN = "-"


def _clear(value):
    """True when the cell explicitly asks for the stored value to be erased."""
    return value.strip() == CLEAR_TOKEN


def _merge(new, current):
    """
    What to store for a field on an existing student.

    blank  -> keep what is there      (so partial sheets are safe)
    "-"    -> erase
    other  -> the new value
    """
    if not new:
        return current
    return "" if _clear(new) else new

PHONE_RE = re.compile(r"^\+?\d{7,15}$")


def clean_phone(raw):
    """Strip formatting and validate. Returns (cleaned, error_or_None)."""
    if not raw:
        return "", "is blank"
    text = re.sub(r"[\s\-().]", "", str(raw).strip())
    if text.startswith("00"):
        text = "+" + text[2:]
    if not PHONE_RE.match(text):
        return "", f"'{raw}' is not a valid phone number"
    return text, None

SPLIT_RE = re.compile(r"[,;|/]+")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[a-zA-Z]{2,}$")


def _canonical(header):
    key = re.sub(r"\s+", " ", str(header or "")).strip().lower().rstrip(":")
    for canon, aliases in HEADER_ALIASES.items():
        if key in aliases:
            return canon
    return None


def read_rows(uploaded_file):
    """Return (list_of_dicts, error_message)."""
    name = (uploaded_file.name or "").lower()
    if name.endswith(".csv"):
        return _read_csv(uploaded_file)
    if name.endswith((".xlsx", ".xlsm")):
        return _read_xlsx(uploaded_file)
    return [], "Unsupported file type. Please upload a .xlsx or .csv file."


def _read_csv(f):
    try:
        text = f.read().decode("utf-8-sig", errors="replace")
    except Exception:
        return [], "Could not read the CSV file."
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], "The file is empty."
    return _normalise(rows[0], rows[1:])


def _read_xlsx(f):
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        return [], "openpyxl is not installed on the server."
    try:
        wb = load_workbook(f, data_only=True, read_only=True)
    except Exception:
        return [], "Could not open the workbook. Is it a valid .xlsx file?"
    ws = wb[wb.sheetnames[0]]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    grid = [r for r in grid if any(c is not None and str(c).strip() for c in r)]
    if not grid:
        return [], "The first worksheet is empty."
    return _normalise(grid[0], grid[1:])


def _normalise(header_row, body):
    mapping = {}
    for idx, cell in enumerate(header_row):
        canon = _canonical(cell)
        if canon and canon not in mapping:
            mapping[canon] = idx
    missing = [c for c in REQUIRED_COLUMNS if c not in mapping]
    if missing:
        return [], (
            "Missing required column(s): %s. Expected headers: email, name, "
            "class roll, batch, subjects enrolled, guardian mobile "
            "(exam roll, mobile number, guardian name and guardian email are optional)."
            % ", ".join(m.replace("_", " ") for m in missing)
        )
    rows = []
    for i, raw in enumerate(body, start=2):
        def cell(key):
            idx = mapping.get(key)
            if idx is None or idx >= len(raw):
                return ""
            val = raw[idx]
            if val is None:
                return ""
            if isinstance(val, float) and val.is_integer():
                val = int(val)
            return str(val).strip()

        rows.append({
            "_row": i,
            "name": cell("name"),
            "email": normalise_email(cell("email")),
            "department": cell("department"),
            "mobile": cell("mobile"),
            "batch": cell("batch"),
            "subjects": cell("subjects"),
            "class_roll": cell("class_roll"),
            "section": cell("section"),
            "exam_roll": cell("exam_roll"),
            "guardian_name": cell("guardian_name"),
            "guardian_mobile": cell("guardian_mobile"),
            "guardian_email": normalise_email(cell("guardian_email")),
            "institute": cell("institute"),
            "discipline": cell("discipline"),
        })
    return rows, None


def _split_subjects(value):
    return [s.strip() for s in SPLIT_RE.split(value or "") if s.strip()]


@transaction.atomic
def import_students(rows, institute, uploader, file_name="roster.xlsx",
                    create_missing_batches=False, send_invites=True,
                    allowed_departments=None):
    """
    Validate & persist a parsed roster. Returns an ImportJob with a per-row
    report. Nothing is written unless the whole file parses (atomic).

    **The department comes from each row, not from the caller.** It used to be
    a single argument chosen from a dropdown, which meant a college with six
    departments uploaded six files and a head could not hand one spreadsheet to
    the registry. `institute` is the scope now, and every row names its own
    department by code or name.

    `allowed_departments` narrows that further and is not optional in practice:
    a HoD may upload, and without it a Department column would let them file
    students into a department they do not run. `None` means "no narrowing" and
    is for callers that have already scoped — the tests, and the seeder.
    """
    from .models import Department

    departments = Department.objects.filter(institute=institute)
    if allowed_departments is not None:
        allowed_ids = {d.pk for d in allowed_departments}
        departments = departments.filter(pk__in=allowed_ids)
    # Code *and* name, lower-cased, the way subjects are matched — a registrar's
    # export spells it whichever way their other system does.
    department_index = {}
    for d in departments:
        department_index[d.code.strip().lower()] = d
        department_index[d.name.strip().lower()] = d

    # Per department, built lazily. One file can now touch every department in
    # the college, and loading every subject and batch up front would be a
    # large query for a sheet that turns out to name one of them.
    subject_indexes = {}
    batch_caches = {}

    def subjects_of(dept):
        if dept.pk not in subject_indexes:
            index = {}
            for s in Subject.objects.filter(department=dept):
                index[s.code.strip().lower()] = s
                index[s.name.strip().lower()] = s
            subject_indexes[dept.pk] = index
        return subject_indexes[dept.pk]

    def batches_of(dept):
        if dept.pk not in batch_caches:
            batch_caches[dept.pk] = {
                b.label.lower(): b for b in Batch.objects.filter(department=dept)}
        return batch_caches[dept.pk]

    created = updated = errors = 0
    report = []
    seen_emails = set()
    # Named, not just counted. "3 sections created" is reassuring; "2022-26 · AA"
    # is the line that tells somebody they typed the section name twice.
    sections_created = []

    # One query for the whole file: which of these emails already exist, and
    # what do they currently hold? Needed both to decide which columns are
    # mandatory and to preserve values the sheet leaves blank.
    existing = {
        p.user.email: p
        for p in StudentProfile.objects.filter(
            user__email__in=[r["email"] for r in rows if r["email"]]
        ).select_related("user", "batch", "department")
    }
    # Departments a row put a student into, for the job record. A file can span
    # several now, so `ImportJob` records the institute and names the
    # departments rather than pretending there was one.
    touched_departments = set()

    for row in rows:
        line, problems = row["_row"], []
        name, email = row["name"], row["email"]
        current = existing.get(email)
        is_new = current is None

        if not email:
            problems.append("email is blank")
        elif not EMAIL_RE.match(email):
            problems.append(f"'{email}' is not a valid email")
        elif email in seen_emails:
            problems.append(f"duplicate email '{email}' in this file")

        # Required columns apply to students being created. For an existing
        # student a blank cell means "not changing this", so demanding a value
        # would defeat the point of a partial update sheet.
        if is_new:
            for field in NEW_STUDENT_REQUIRED:
                if not row.get(field):
                    problems.append(
                        f"{field.replace('_', ' ')} is blank — required for a new student")

        # The department this row's student belongs to. Resolved before
        # anything that depends on it — subjects and batches are looked up
        # inside it, so a row that names an unknown department cannot be
        # checked any further and says so once rather than three times.
        department = None
        if row.get("department") and not _clear(row["department"]):
            department = department_index.get(row["department"].strip().lower())
            if department is None:
                problems.append(
                    f"department '{row['department']}' is not one you manage — "
                    f"check the spelling, or use its code")
        elif current is not None:
            # Blank on an existing student keeps theirs, like every other
            # column. (A new student with no department was rejected above.)
            department = current.department

        # The institute, if the sheet supplied it. Compared, never applied — a
        # student's institute follows from their department. Discipline used to
        # be checked here too and is not any more: it follows from the
        # department, which the row now names, so there is nothing left for it
        # to catch.
        if row.get("institute") and not _clear(row["institute"]):
            if not _same(row["institute"], institute.name):
                problems.append(
                    f"institute '{row['institute']}' is not "
                    f"{institute.name} — is this the right file?")

        parsed = None
        if row["batch"]:
            parsed = parse_batch_label(row["batch"])
            if parsed is None:
                problems.append(f"batch '{row['batch']}' is not in the format 2022-26")

        # A guardian mobile is only validated when one is supplied; blank on an
        # existing student simply leaves the stored number in place.
        guardian_mobile = ""
        if row["guardian_mobile"] and not _clear(row["guardian_mobile"]):
            guardian_mobile, phone_error = clean_phone(row["guardian_mobile"])
            if phone_error:
                problems.append(f"guardian mobile {phone_error}")
        student_mobile = ""
        if row["mobile"] and not _clear(row["mobile"]):
            student_mobile, _ = clean_phone(row["mobile"])

        # Subjects are looked up *inside this row's department*, so "DSA" in
        # one department and "DSA" in another are two different papers — which
        # they are. Skipped entirely when the department did not resolve:
        # every code would be reported unknown, burying the one problem that
        # matters under a list of consequences.
        subject_names = _split_subjects(row["subjects"])
        matched, unknown = [], []
        if department is not None:
            index = subjects_of(department)
            for token in subject_names:
                hit = index.get(token.lower())
                (matched.append(hit) if hit else unknown.append(token))
            if unknown:
                problems.append(
                    "unknown subject(s) in %s: %s — add them under Subjects first"
                    % (department.code, ", ".join(unknown))
                )

        if problems:
            errors += 1
            report.append({"row": line, "email": email, "status": "error",
                           "department": row.get("department", ""),
                           "institute": row.get("institute", ""),
                           "messages": problems})
            continue

        seen_emails.add(email)
        touched_departments.add(department.code)
        batch_cache = batches_of(department)

        if parsed is None:
            # No batch column for a student who already exists: keep theirs.
            # (A new student without a batch was already rejected above.)
            batch, label = current.batch, current.batch.label
            # …unless the row also moved them to another department, in which
            # case their old batch belongs to the department they are leaving.
            # Refusing beats guessing: a batch is a cohort, and picking one for
            # somebody would put them in a year nobody chose.
            if batch.department_id != department.pk:
                errors += 1
                report.append({
                    "row": line, "email": email, "status": "error",
                    "messages": [
                        f"this row moves {email} to {department.code}, so it "
                        f"has to name a batch in {department.code} — "
                        f"{label} belongs to {batch.department.code}"],
                })
                continue
        else:
            start, end, label = parsed
            batch = batch_cache.get(label.lower())
        if batch is None:
            # Creating batches from a spreadsheet used to be the default, which
            # meant a typo — 2022-27 for 2022-26 — silently produced a second
            # cohort and split a class in two. Nobody sees that until the
            # attendance figures stop adding up.
            if not create_missing_batches:
                errors += 1
                report.append({
                    "row": line, "email": email, "status": "error",
                    "messages": [f"batch {label} does not exist — create it under "
                                 "Batches first, or correct the spelling"],
                })
                continue
            batch = Batch.objects.create(
                department=department, label=label, start_year=start, end_year=end
            )
            batch_cache[label.lower()] = batch
        elif not batch.is_active:
            # Importing into an archived batch would create students that are
            # invisible everywhere — refuse rather than confuse.
            errors += 1
            report.append({
                "row": line, "email": email, "status": "error",
                "messages": [f"batch {label} is archived — restore it under "
                             "Batches before importing into it"],
            })
            continue

        user, invitation, was_created = invite_user(
            email=email,
            role="STUDENT",
            institute=department.institute,
            department=department,
            full_name=name or (current.user.full_name if current else ""),
            invited_by=uploader,
            payload={"batch": label, "class_roll": row["class_roll"],
                     "exam_roll": row["exam_roll"]},
            extra_lines=[
                f"Department: {department.name} ({department.code})",
                f"Batch: {label}",
                "Subjects: " + ", ".join(s.code for s in matched),
            ],
            # Never from inside: whether to email is decided below, on whether
            # this row produced a new account.
            send=False,
        )
        if invitation is None and user.registration_completed:
            # Existing active student — just refresh their enrolment data.
            was_created = False

        # Only a genuinely new account gets an invitation. A sheet of email +
        # guardian mobile is a routine bulk edit, and it used to re-mail an
        # invitation link to every student who had not yet activated — every
        # time it was uploaded. An address that changed arrives here as a new
        # account anyway, so it is covered by the same rule.
        invited = bool(send_invites and was_created and invitation is not None)
        if invited:
            send_invitation(invitation, extra_lines=[
                f"Department: {department.name} ({department.code})",
                f"Batch: {label}",
                "Subjects: " + ", ".join(s.code for s in matched),
            ])

        # The section, created if this batch has not seen it before. The sheet
        # is the source of truth for a new intake, so an import of 200 students
        # across A–D produces the four sections rather than 200 errors. How
        # many were created is reported at the end, which is what makes a typo
        # like "AA" visible instead of silent.
        #
        # Resolved against `batch` — the row's batch, not the student's current
        # one — so a row that moves somebody to another cohort puts them in a
        # section of the cohort they are moving *to*.
        section = current.section if current else None
        if row["section"] and not _clear(row["section"]):
            section, made = sections.resolve(batch, row["section"], create=True)
            if made:
                sections_created.append(f"{batch.label} · {section.name}")
        elif _clear(row["section"]):
            section = None
        # A student kept from a previous import may hold a section of the batch
        # they are leaving. Clearing beats carrying it across: an unsectioned
        # student is an ordinary state, one listed under another cohort's
        # section is a quiet error in two rosters.
        if section is not None and section.batch_id != batch.pk:
            section = None

        # Blank cell = leave it alone. Without this a sheet of just
        # email + guardian mobile would erase every other field on the row.
        profile, p_created = StudentProfile.objects.update_or_create(
            user=user,
            defaults={
                "department": department,
                "batch": batch,
                "section": section,
                "class_roll": _merge(row["class_roll"], current.class_roll if current else ""),
                "exam_roll": _merge(row["exam_roll"], current.exam_roll if current else ""),
                "mobile": _merge(student_mobile or row["mobile"],
                                 current.mobile if current else ""),
                "guardian_name": _merge(row["guardian_name"],
                                        current.guardian_name if current else ""),
                "guardian_mobile": _merge(guardian_mobile or row["guardian_mobile"],
                                          current.guardian_mobile if current else ""),
                "guardian_email": _merge(row["guardian_email"],
                                         current.guardian_email if current else ""),
                "is_active": True,
            },
        )
        if student_mobile and not user.phone:
            user.phone = student_mobile
            user.save(update_fields=["phone"])

        # An empty subjects column means "not touching enrolments" for an
        # existing student, so only rewrite them when the sheet says something.
        if matched:
            wanted = {s.id for s in matched}
            Enrollment.objects.filter(student=profile).exclude(
                subject_id__in=wanted).update(is_active=False)
            for subj in matched:
                Enrollment.objects.update_or_create(
                    student=profile, subject=subj, defaults={"is_active": True}
                )

        if was_created or p_created:
            created += 1
            status = "created"
        else:
            updated += 1
            status = "updated"
        report.append({
            "row": line, "email": email, "name": profile.name, "batch": label,
            "section": profile.section.name if profile.section_id else "",
            "class_roll": profile.class_roll, "exam_roll": profile.exam_roll,
            "guardian_mobile": profile.guardian_mobile,
            "subjects": [s.code for s in matched] or
                        [e.subject.code for e in profile.enrollments.filter(is_active=True)],
            # Shown in the preview so "who will be emailed" is visible before
            # anything is sent, not discovered afterwards.
            "invited": invited,
            # Echoed from the *resolved* department rather than from the cell.
            # "CSE" in the sheet and "Computer Science" on screen is the row
            # working; the point of showing it is to confirm which department a
            # row actually landed in, which the cell cannot say.
            "department": f"{department.name} ({department.code})",
            "institute": institute.name,
            "status": status, "messages": [],
        })

    if errors and (created or updated):
        job_status = ImportJob.Status.PARTIAL
    elif errors:
        job_status = ImportJob.Status.FAILED
    else:
        job_status = ImportJob.Status.SUCCESS

    return ImportJob.objects.create(
        # The institute, not a department. A file can name several now, so
        # picking one for the record would be picking a winner — the ones it
        # actually touched are listed in the report instead.
        institute=institute,
        uploaded_by=uploader,
        file_name=file_name[:255],
        status=job_status,
        total_rows=len(rows),
        created_count=created,
        updated_count=updated,
        error_count=errors,
        # Listed by name so a mistyped section reads as a mistake rather than
        # as a count. Kept beside the rows rather than inside them: they are
        # facts about the file, not about any one student.
        report={"rows": report, "sections_created": sections_created,
                "departments": sorted(touched_departments)},
    )


def build_template_workbook(department=None):
    """Return an in-memory .xlsx the HoD can download as a starting point."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    headers = list(ROSTER_COLUMNS)
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="1F3B73")
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 26
    ws.freeze_panes = "A2"

    codes = "DSA, DBMS, AI"
    dept_code = "CSE"
    if department is not None:
        dept_code = department.code
        subs = list(Subject.objects.filter(department=department, is_active=True)[:3])
        if subs:
            codes = ", ".join(s.code for s in subs)
    institute_name = department.institute.name if department is not None else ""
    # Two sample rows, deliberately in different sections: one example of a
    # column teaches the format, two teach that it varies per student. Both
    # carry the department code, because that column is an instruction now —
    # the file decides where each student goes, not a dropdown on the modal.
    ws.append(["ananya@example.com", "Ananya Sharma", dept_code, "01", "CSE22001",
               "2022-26", "A", codes,
               "+919812345670", "Mr. R. Sharma", "sharma@example.com", "9876543210",
               institute_name])
    ws.append(["rahul@example.com", "Rahul Verma", dept_code, "02", "CSE22002",
               "2022-26", "B", codes,
               "+919812345671", "Mrs. S. Verma", "", "9876500011",
               institute_name])

    notes = wb.create_sheet("Instructions")
    for line in [
        ["Column", "Required", "Notes"],
        ["Email", "Always", "The key this sheet matches on. An email that already "
                            "exists updates that student instead of creating one."],
        ["Name", "New students", "Student's full name"],
        ["Department", "New students", "Code or name, e.g. CSE or Computer "
                                       "Science. One file can carry every "
                                       "department in the institute — each row "
                                       "says where its student belongs. Blank "
                                       "on an existing student keeps theirs."],
        ["Class Roll", "New students", "Day-to-day roll inside the batch, e.g. 01"],
        ["Exam Roll", "No", "University / registration number, e.g. CSE22001"],
        ["Batch", "New students", "Format 2022-26 (created automatically if new)"],
        ["Section", "No", "Which section of the batch, e.g. A. Created "
                          "automatically if this batch has not used it before, "
                          "so check the spelling — 'AA' makes a section called "
                          "AA rather than an error. Leave blank for none."],
        ["Subjects Enrolled", "New students", "Comma separated subject codes or names"],
        ["Guardian Mobile", "New students", "WhatsApp number for low-attendance alerts"],
        ["Guardian Name", "No", ""],
        ["Guardian Email", "No", ""],
        ["Mobile Number", "No", "10-digit mobile"],
        ["Institute", "No", "Checked, not applied. A student's institute follows "
                            "from the department their row names; this column "
                            "only catches a sheet uploaded to the wrong account."],
        ["", "", ""],
        ["UPDATING EXISTING STUDENTS", "", ""],
        ["Blank cell", "", "Leaves the stored value unchanged — so a sheet of just "
                           "Email + Guardian Mobile updates only guardian mobiles."],
        ["A single -", "", "Clears the stored value."],
        ["Subjects blank", "", "Enrolments are left exactly as they are."],
    ]:
        notes.append(line)
    for col in range(1, 4):
        notes.column_dimensions[get_column_letter(col)].width = 34
        notes.cell(row=1, column=col).font = Font(bold=True)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def build_roster_workbook(students):
    """
    Export the current roster in exactly the shape `import_students` accepts,
    so an export can be edited and re-imported round-trip.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    headers = ROSTER_COLUMNS + EXPORT_ONLY_COLUMNS
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F3B73")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 24
    ws.freeze_panes = "A2"

    for student in students:
        ws.append([
            student.user.email,
            student.user.full_name,
            # The code, not the name: it is what the template's sample rows
            # show, it is shorter to retype, and both resolve on import anyway.
            student.department.code,
            student.class_roll,
            student.exam_roll,
            student.batch.label,
            # In the same slot the template puts it, so a round trip lines up
            # cell for cell — the one list `ROSTER_COLUMNS` exists to guarantee.
            student.section.name if student.section_id else "",
            ", ".join(e.subject.code for e in student.enrollments.all() if e.is_active),
            student.guardian_mobile,
            student.guardian_name,
            student.guardian_email,
            student.mobile or student.user.phone,
            # Written out on export and checked-not-applied on import — see the
            # note beside its header alias. Discipline used to sit beside it and
            # is gone: it follows from the department, which the row now names.
            student.department.institute.name,
            "Activated" if student.user.registration_completed else "Invited",
        ])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
