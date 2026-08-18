"""
Discipline on a department, and what follows from it.

Three claims:

* every row a Subjects / Batches / Teachers / Students table renders carries an
  institute and a discipline, for every kind of user rather than only a
  university;
* the student import checks those two columns against the department it is
  loading into, and never applies them;
* a subject or batch in an *affiliated* department is read-only for the
  institute, and one in an autonomous department is not.

The third is the behaviour change. The rule used to lock only rows a university
had pushed, which left a gap: an affiliated institute could add its own subject
beside them and edit it freely.
"""
import json

from django.test import RequestFactory, TestCase
from django.urls import reverse

from academics import views
from academics.curriculum import governing_university, is_read_only
from academics.models import Batch, Department, StudentProfile, Subject
from accounts.models import (
    Discipline,
    Institute,
    InstituteAffiliation,
    University,
    User,
)


class DisciplineFixture(TestCase):
    def setUp(self):
        self.university = University.objects.create(
            name="Engineering University", code="ENGGU", short_name="ENGGU",
            email="e@u.edu", grants_affiliation=True)
        self.institute = Institute.objects.create(
            name="Acme College", code="ACME", email="office@acme.edu",
            status=Institute.Status.APPROVED)
        # Engineering is theirs; general courses are ours.
        InstituteAffiliation.objects.create(
            institute=self.institute, discipline=Discipline.ENGG,
            university=self.university)
        InstituteAffiliation.objects.create(
            institute=self.institute, discipline=Discipline.GENERAL,
            university=None)

        from academics.models import UniversityDepartment

        entry = UniversityDepartment.objects.create(
            university=self.university, discipline=Discipline.ENGG,
            name="Computer Science", code="CSE")
        # Adopted, so its rows are the university's — see academics/catalogue.py.
        self.cse = Department.objects.create(
            institute=self.institute, code="CSE", name="Computer Science",
            discipline=Discipline.ENGG, source=entry)
        self.arts = Department.objects.create(
            institute=self.institute, code="ARTS", name="Arts",
            discipline=Discipline.GENERAL)
        self.legacy = Department.objects.create(
            institute=self.institute, code="OLD", name="Legacy", discipline="")

        self.head = User.objects.create_user(
            email="head@acme.edu", password="Str0ngPass!23",
            role=User.Role.HEAD, institute=self.institute,
            registration_completed=True)

    def _rows(self, url_name, user, **params):
        self.client.force_login(user)
        response = self.client.get(reverse(url_name), params,
                                   HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        return response.json()["data"]["rows"]


class GoverningTests(DisciplineFixture):
    def test_an_affiliated_department_is_governed_by_its_university(self):
        self.assertEqual(governing_university(self.cse), self.university)

    def test_an_autonomous_department_is_governed_by_nobody(self):
        self.assertIsNone(governing_university(self.arts))

    def test_a_department_with_no_discipline_is_governed_by_nobody(self):
        """Every department that predates the column."""
        self.assertIsNone(governing_university(self.legacy))


class ReadOnlyTests(DisciplineFixture):
    def _subject(self, department, code):
        return Subject.objects.create(
            department=department, code=code, name=code,
            semester=1, credits=3)

    def test_an_adopted_subject_is_read_only(self):
        subject = self._subject(self.cse, "DSA")
        from academics.models import UniversitySubject

        subject.source = UniversitySubject.objects.create(
            department=self.cse.source, code="DSA", name="DSA")
        subject.save()
        self.assertTrue(is_read_only(subject, self.head))

    def test_a_subject_in_an_autonomous_department_is_not(self):
        self.assertFalse(is_read_only(self._subject(self.arts, "ENG"), self.head))

    def test_a_batch_follows_the_same_rule_as_a_subject(self):
        from academics.models import UniversityBatch

        published = UniversityBatch.objects.create(
            department=self.cse.source, label="2022-26",
            start_year=2022, end_year=2026)
        locked = Batch.objects.create(department=self.cse, label="2022-26",
                                      start_year=2022, end_year=2026,
                                      source=published)
        free = Batch.objects.create(department=self.arts, label="2022-25",
                                    start_year=2022, end_year=2025)
        self.assertTrue(is_read_only(locked, self.head))
        self.assertFalse(is_read_only(free, self.head))

    def test_the_governing_university_is_never_locked_out(self):
        admin = User.objects.create_user(
            email="admin@u.edu", password="Str0ngPass!23",
            role=User.Role.UNIVERSITY, university=self.university)
        self.assertFalse(is_read_only(self._subject(self.cse, "DSA"), admin))

    def test_a_subject_the_institute_added_to_an_adopted_department_is_theirs(self):
        """
        Changed with the catalogue, and deliberately.

        The department is the university's; a paper the college added alongside
        it is not. The old rule locked everything in an affiliated department,
        which meant a college could not add so much as an elective. `source`
        draws the line where it actually falls.
        """
        mine = self._subject(self.cse, "MINE")
        self.assertIsNone(mine.source_id)
        self.assertFalse(is_read_only(mine, self.head))

    def test_the_endpoint_refuses_the_edit_not_just_the_button(self):
        from academics.models import UniversitySubject

        subject = self._subject(self.cse, "DSA")
        subject.source = UniversitySubject.objects.create(
            department=self.cse.source, code="DSA", name="DSA")
        subject.save()
        request = RequestFactory().post("/", {
            "code": "DSA", "name": "Renamed", "semester": 1, "credits": 3,
            "subject_type": "THEORY", "degree": "BACHELOR", "is_active": "on"})
        request.user = self.head
        request.session = self.client.session
        request.META["HTTP_X_REQUESTED_WITH"] = "XMLHttpRequest"
        response = views.api_subject_save(request, pk=subject.pk)
        self.assertEqual(response.status_code, 403)


class ColumnPayloadTests(DisciplineFixture):
    """The two columns reach the client on all four tables."""

    def setUp(self):
        super().setUp()
        Subject.objects.create(department=self.cse, code="DSA",
                               name="Data Structures", semester=1, credits=3)
        self.batch = Batch.objects.create(
            department=self.cse, label="2022-26",
            start_year=2022, end_year=2026)
        User.objects.create_user(
            email="t@acme.edu", password="Str0ngPass!23", full_name="A Teacher",
            role=User.Role.TEACHER, institute=self.institute,
            department=self.cse, registration_completed=True)
        student = User.objects.create_user(
            email="s@acme.edu", password="Str0ngPass!23", full_name="A Student",
            role=User.Role.STUDENT, institute=self.institute)
        StudentProfile.objects.create(
            user=student, department=self.cse, batch=self.batch,
            class_roll="1")

    def _assert_both(self, rows):
        self.assertTrue(rows)
        for row in rows:
            self.assertIn("institute", row)
            self.assertIn("discipline_label", row)
            # The column shows the code; the full name travels beside it for
            # the tooltip. Both, or the tooltip silently shows nothing.
            self.assertIn("institute_name", row)

    def test_subjects_carry_both(self):
        rows = self._rows("academics:api_subjects", self.head)
        self._assert_both(rows)
        self.assertEqual(rows[0]["discipline"], "ENGG")
        self.assertEqual(rows[0]["institute"], "ACME")
        self.assertEqual(rows[0]["institute_name"], "Acme College")

    def test_batches_carry_both(self):
        self._assert_both(self._rows("academics:api_batches", self.head))

    def test_teachers_carry_both(self):
        self._assert_both(self._rows("academics:api_teachers", self.head))

    def test_students_carry_both(self):
        self._assert_both(self._rows("academics:api_students", self.head))

    def test_the_institute_column_carries_the_code_on_every_one_of_them(self):
        """
        Item 1. The column is an identifier, not a label — a full institute
        name wraps and pushes a dense table sideways.
        """
        for url in ("academics:api_subjects", "academics:api_batches",
                    "academics:api_teachers", "academics:api_students",
                    "academics:api_departments"):
            with self.subTest(url=url):
                rows = self._rows(url, self.head)
                self.assertEqual(rows[0]["institute"], "ACME", url)
                self.assertEqual(rows[0]["institute_name"], "Acme College", url)

    def test_departments_carry_the_discipline_they_define(self):
        rows = self._rows("academics:api_departments", self.head)
        by_code = {r["code"]: r for r in rows}
        self.assertEqual(by_code["CSE"]["discipline"], "ENGG")
        self.assertEqual(by_code["OLD"]["discipline"], "")

    def test_a_teacher_sees_the_columns_too_not_only_a_university(self):
        """
        The requirement is "for every type of users". The Institute column is
        university-only everywhere else in the app; on these four tables it is
        not, and the payload has to carry it regardless of who asks.
        """
        teacher = User.objects.get(email="t@acme.edu")
        # The student directory rather than Subjects: a teacher's subject list
        # is scoped to what they are allocated to teach, and this one has no
        # allocations, so an empty list there would prove nothing either way.
        rows = self._rows("academics:api_students", teacher, scope="all")
        self._assert_both(rows)
        self.assertEqual(rows[0]["institute"], "ACME")
        self.assertEqual(rows[0]["institute_name"], "Acme College")
        self.assertEqual(rows[0]["discipline_label"],
                         "Engineering, Technology & Management")


class ImportColumnTests(DisciplineFixture):
    """
    Institute and Discipline in the sheet are checks, not instructions.

    A cell cannot move a student to another institute — that follows from the
    department being imported into. What it can do is catch a sheet loaded into
    the wrong department, which is the mistake actually worth catching.
    """

    def _run(self, rows, department=None):
        """
        Parse a sheet and run the import in dry-run, returning the report.

        Goes through `_normalise` rather than a hand-built dict so the header
        aliases are exercised too — "Institute" has to be recognised as the
        institute column before anything can be checked against it.
        """
        from academics.importer import _normalise, import_students

        header = ["Email", "Name", "Class Roll", "Batch", "Subjects Enrolled",
                  "Guardian Mobile", "Institute", "Discipline"]
        parsed, error = _normalise(header, [list(r) for r in rows])
        self.assertIsNone(error, error)
        job = import_students(parsed, department or self.cse, self.head,
                              create_missing_batches=True, send_invites=False)
        return job.report          # already {"rows": [...]}

    def test_a_matching_institute_and_discipline_pass(self):
        Subject.objects.create(department=self.cse, code="DSA", name="DSA",
                               semester=1, credits=3)
        result = self._run([[
            "a@acme.edu", "A", "01", "2022-26", "DSA", "+919812345670",
            "Acme College", "Engineering, Technology & Management"]])
        messages = [m for r in result["rows"] for m in r.get("messages", [])]
        self.assertFalse([m for m in messages if "institute" in m or "discipline" in m],
                         messages)

    def test_the_wrong_institute_is_caught_before_anything_lands(self):
        result = self._run([[
            "a@acme.edu", "A", "01", "2022-26", "DSA", "+919812345670",
            "Some Other College", ""]])
        messages = " ".join(m for r in result["rows"] for m in r.get("messages", []))
        self.assertIn("is not Acme College", messages)

    def test_the_wrong_discipline_is_caught(self):
        result = self._run([[
            "a@acme.edu", "A", "01", "2022-26", "DSA", "+919812345670",
            "", "Pharmacy"]])
        messages = " ".join(m for r in result["rows"] for m in r.get("messages", []))
        self.assertIn("discipline 'Pharmacy' is not", messages)

    def test_the_stored_code_is_accepted_as_well_as_the_label(self):
        Subject.objects.create(department=self.cse, code="DSA", name="DSA",
                               semester=1, credits=3)
        result = self._run([[
            "a@acme.edu", "A", "01", "2022-26", "DSA", "+919812345670",
            "", "ENGG"]])
        messages = " ".join(m for r in result["rows"] for m in r.get("messages", []))
        self.assertNotIn("discipline", messages)

    def test_omitting_both_columns_checks_nothing(self):
        """They are optional; a sheet without them behaves exactly as before."""
        Subject.objects.create(department=self.cse, code="DSA", name="DSA",
                               semester=1, credits=3)
        result = self._run([[
            "a@acme.edu", "A", "01", "2022-26", "DSA", "+919812345670", "", ""]])
        messages = " ".join(m for r in result["rows"] for m in r.get("messages", []))
        self.assertNotIn("institute", messages)

    def test_a_discipline_named_for_a_department_that_has_none_is_flagged(self):
        result = self._run([[
            "a@acme.edu", "A", "01", "2022-26", "DSA", "+919812345670",
            "", "ENGG"]], department=self.legacy)
        messages = " ".join(m for r in result["rows"] for m in r.get("messages", []))
        self.assertIn("has none on file", messages)


class RosterColumnOrderTests(DisciplineFixture):
    """
    The export and the import template must line up cell for cell.

    They were two literal header lists, and the moment the template gained
    Institute and Discipline the export still ended at Mobile Number — so a
    round trip put values under the wrong headings. One shared list cannot
    drift from itself; these tests are what keep it shared.
    """

    def _headers(self, stream):
        from openpyxl import load_workbook

        return [c.value for c in load_workbook(stream).active[1]]

    def test_the_export_starts_with_exactly_the_template_columns(self):
        from academics.importer import (
            ROSTER_COLUMNS,
            build_roster_workbook,
            build_template_workbook,
        )

        template = self._headers(build_template_workbook(self.cse))
        export = self._headers(build_roster_workbook([]))
        self.assertEqual(template, ROSTER_COLUMNS)
        self.assertEqual(export[:len(ROSTER_COLUMNS)], ROSTER_COLUMNS)

    def test_the_export_only_extras_come_after_the_shared_ones(self):
        from academics.importer import (
            EXPORT_ONLY_COLUMNS,
            ROSTER_COLUMNS,
            build_roster_workbook,
        )

        export = self._headers(build_roster_workbook([]))
        self.assertEqual(export, ROSTER_COLUMNS + EXPORT_ONLY_COLUMNS)

    def test_every_exported_row_is_as_wide_as_the_header(self):
        """
        A row shorter than its header silently shifts nothing; a row *longer*
        shifts everything after it. Either way the fix is to check the width.
        """
        from openpyxl import load_workbook

        from academics.importer import build_roster_workbook
        from academics.models import Batch, StudentProfile

        batch = Batch.objects.create(department=self.cse, label="2022-26",
                                     start_year=2022, end_year=2026)
        user = User.objects.create_user(
            email="s@acme.edu", password="Str0ngPass!23", full_name="A Student",
            role=User.Role.STUDENT, institute=self.institute)
        profile = StudentProfile.objects.create(
            user=user, department=self.cse, batch=batch, class_roll="1")

        sheet = load_workbook(build_roster_workbook([profile])).active
        header = [c.value for c in sheet[1]]
        row = [c.value for c in sheet[2]]
        self.assertEqual(len(row), len(header))
        # And the two new cells carry what they claim to.
        self.assertEqual(row[header.index("Institute")], "Acme College")
        self.assertEqual(row[header.index("Discipline")],
                         "Engineering, Technology & Management")

    def test_an_exported_sheet_re_imports_without_column_complaints(self):
        """The round trip the shared list exists to protect."""
        from openpyxl import load_workbook

        from academics.importer import _normalise, build_roster_workbook
        from academics.models import Batch, StudentProfile

        batch = Batch.objects.create(department=self.cse, label="2022-26",
                                     start_year=2022, end_year=2026)
        user = User.objects.create_user(
            email="s@acme.edu", password="Str0ngPass!23", full_name="A Student",
            role=User.Role.STUDENT, institute=self.institute)
        profile = StudentProfile.objects.create(
            user=user, department=self.cse, batch=batch, class_roll="1")

        sheet = load_workbook(build_roster_workbook([profile])).active
        header = [c.value for c in sheet[1]]
        body = [[c.value for c in r] for r in sheet.iter_rows(min_row=2)]
        rows, error = _normalise(header, body)
        self.assertIsNone(error, error)
        self.assertEqual(rows[0]["email"], "s@acme.edu")
        self.assertEqual(rows[0]["institute"], "Acme College")
